from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import shutil
import os
import tempfile
import pandas as pd
import pytz
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
from pytz import timezone
from detect_count import detect_and_count

app = FastAPI()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:4173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

HISTORY_FILE = 'history.csv'
STATIC_DIR = 'static'

if not os.path.exists(STATIC_DIR):
    os.makedirs(STATIC_DIR)


app.mount("/static", StaticFiles(directory="static"), name="static")

def log_to_csv(record: dict):
    df = pd.DataFrame([record])
    header = not os.path.exists(HISTORY_FILE)
    df.to_csv(HISTORY_FILE, mode='a', header=header, index=False)

@app.post("/count/")
async def count(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    
    output_img_path = os.path.join(STATIC_DIR, f"{os.path.basename(tmp_path).replace(suffix, '')}_out{suffix}")
    result = detect_and_count(tmp_path, output_path=output_img_path)

    record = {
        'timestamp': datetime.now(tz=timezone('utc')).isoformat(),
        'people_total': result['people_total'],
        'tables_counts': str([t['count'] for t in result['tables']])
    }
    log_to_csv(record)

    try:
        os.remove(tmp_path)
    except OSError:
        pass

    return JSONResponse({
        'tables': result['tables'],
        'people_total': result['people_total'],
        'img_url': f'http://localhost:8000/static/{os.path.basename(output_img_path)}'
    })

@app.get("/history/")
def get_history():
    if not os.path.exists(HISTORY_FILE):
        return JSONResponse([], status_code=200)
    df = pd.read_csv(HISTORY_FILE)
    return df.to_dict(orient='records')

@app.get("/export/csv/")
def export_csv():
    if not os.path.exists(HISTORY_FILE):
        return JSONResponse({'error': 'No history'}, status_code=404)
    return FileResponse(HISTORY_FILE, media_type='text/csv', filename=HISTORY_FILE)

@app.get("/export/excel/")
def export_excel():
    if not os.path.exists(HISTORY_FILE):
        return JSONResponse({'error': 'No history'}, status_code=404)
    
    df = pd.read_csv(HISTORY_FILE)

    df['timestamp'] = pd.to_datetime(df['timestamp'])

    if df['timestamp'].dt.tz is None:
        df['timestamp'] = df['timestamp'].dt.tz_localize('UTC')

    msk_tz = pytz.timezone('Europe/Moscow')
    df['timestamp'] = df['timestamp'].dt.tz_convert(msk_tz)

    df['timestamp'] = df['timestamp'].dt.tz_localize(None)

    excel_file = 'history.xlsx'
    df.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    date_style = NamedStyle(name="datetime", number_format='YYYY-MM-DD HH:MM:SS')

    for row in range(2, len(df) + 2):
        cell = ws.cell(row=row, column=1)
        cell.style = date_style

    wb.save(excel_file)

    return FileResponse(excel_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=excel_file)

# Запуск: uvicorn app:app --reload --host 0.0.0.0 --port 8000